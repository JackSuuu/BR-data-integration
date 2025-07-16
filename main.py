import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import shutil
from copy import copy
import re

def remove_external_links(workbook):
    """Remove external links from the workbook to prevent security warnings"""
    print("  🔗 Removing external links to prevent security warnings...")
    
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Check if cell contains external link patterns
                    if re.search(r'\[.*\.xl.*\]', str(cell.value)) or "'" in str(cell.value):
                        # Convert formula to value if it's a formula with external reference
                        try:
                            # If it's a formula, try to keep just the result
                            if str(cell.value).startswith('='):
                                # Remove the formula and keep the last calculated value
                                # For safety, we'll clear external reference formulas
                                if '[' in str(cell.value) and ']' in str(cell.value):
                                    cell.value = None  # Clear external references
                        except Exception:
                            pass

def extract_date_from_sheet(client_ws):
    """Extract date from the client sheet itself"""
    # Look for date in common locations (adjust based on your sheet structure)
    date_value = None
    
    # Check first few rows and columns for date-like values
    for row in range(1, 6):  # Check first 5 rows
        for col in range(1, 6):  # Check first 5 columns
            cell_value = client_ws.cell(row=row, column=col).value
            if cell_value:
                # Check if it's a date
                if isinstance(cell_value, datetime):
                    date_value = cell_value.strftime("%Y%m%d")
                    break
                # Check if it's a string that looks like a date
                elif isinstance(cell_value, str):
                    # Try to parse various date formats
                    for date_format in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y%m%d"]:
                        try:
                            parsed_date = datetime.strptime(cell_value, date_format)
                            date_value = parsed_date.strftime("%Y%m%d")
                            break
                        except ValueError:
                            continue
                    if date_value:
                        break
        if date_value:
            break
    
    return date_value

def copy_tab1_format_with_client_data(source_wb, tab1_ws, client_ws, new_tab_name):
    """Create a new tab with tab1 format but client data"""
    # Create new worksheet
    new_ws = source_wb.create_sheet(title=new_tab_name)
    
    # First, copy all formatting from tab1
    for row in tab1_ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column)
            
            # Copy all formatting
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    
    # Copy column dimensions
    for col in tab1_ws.column_dimensions:
        new_ws.column_dimensions[col].width = tab1_ws.column_dimensions[col].width
    
    # Copy row dimensions
    for row in tab1_ws.row_dimensions:
        new_ws.row_dimensions[row].height = tab1_ws.row_dimensions[row].height
    
    # Copy merged cells
    for merged_cell_range in tab1_ws.merged_cells.ranges:
        new_ws.merge_cells(str(merged_cell_range))
    
    # Now copy client data into the formatted cells
    for row in client_ws.iter_rows():
        for cell in row:
            if cell.value is not None:  # Only copy non-empty cells
                target_cell = new_ws.cell(row=cell.row, column=cell.column)
                target_cell.value = cell.value
                # The formatting is already set above, so we just set the value
    
    return new_ws

def read_account_list():
    """Read the account list to get client information"""
    try:
        account_df = pd.read_excel('account_list.xlsx')
        print("account_list.xlsx loaded successfully")
        
        # Get list of clients
        if 'Client' in account_df.columns:
            clients = account_df['Client'].dropna().unique().tolist()
            print(f"Found clients in account list: {clients}")
            return clients
        else:
            print("'Client' column not found in account_list.xlsx")
            print("Available columns:", account_df.columns.tolist())
            return []
    except FileNotFoundError:
        print("account_list.xlsx not found in current directory")
        return []
    except Exception as e:
        print(f"Error reading account_list.xlsx: {e}")
        return []

def get_client_files():
    """Get all Excel files from client_portfolio folder"""
    client_files = []
    portfolio_path = 'client_portfolio'
    
    if not os.path.exists(portfolio_path):
        print(f"Directory {portfolio_path} does not exist")
        return client_files
    
    # Get all Excel files in the client_portfolio folder
    excel_patterns = ['*.xlsx', '*.xls', '*.xlsm']
    for pattern in excel_patterns:
        files = glob.glob(os.path.join(portfolio_path, pattern))
        # Filter out temporary Excel files (starting with ~$)
        files = [f for f in files if not os.path.basename(f).startswith('~$')]
        client_files.extend(files)
    
    print(f"Found {len(client_files)} client files: {client_files}")
    return client_files

def extract_client_name(filename):
    """Extract client name from filename (assumes format: client_name_date.xlsx)"""
    basename = os.path.basename(filename)
    # Remove file extension
    name_without_ext = os.path.splitext(basename)[0]
    # Split by underscore and take the first part as client name
    parts = name_without_ext.split('_')
    if len(parts) >= 2:
        return '_'.join(parts[:-1])  # Everything except the last part (date)
    return name_without_ext

def extract_date_from_filename(filename):
    """Extract date from filename for sorting"""
    basename = os.path.basename(filename)
    name_without_ext = os.path.splitext(basename)[0]
    parts = name_without_ext.split('_')
    if len(parts) >= 2:
        return parts[-1]  # Last part should be the date
    return ""

def create_summary_for_client(client_name, client_files_list):
    """Create a summary file for a specific client"""
    print(f"\n🔄 Creating summary for client: {client_name}")
    
    # Create output filename for this client
    output_filename = f'{client_name}_summary.xlsx'
    
    # Copy Template.xlsx as the starting point
    print(f"  Copying Template.xlsx as base for {output_filename}...")
    shutil.copy2('Template.xlsx', output_filename)
    
    # Load the copied workbook
    summary_wb = load_workbook(output_filename)
    
    # Remove external links to prevent security warnings
    remove_external_links(summary_wb)
    
    # Remove unwanted sheets
    sheets_to_remove = ['Calculations']
    for sheet_name in sheets_to_remove:
        if sheet_name in summary_wb.sheetnames:
            summary_wb.remove(summary_wb[sheet_name])
            print(f"  Removed sheet: {sheet_name}")
    
    # Check if tab1 exists in the template
    if 'tab1' not in summary_wb.sheetnames:
        print(f"  Error: 'tab1' sheet not found in Template.xlsx")
        print(f"  Available sheets: {summary_wb.sheetnames}")
        return False
    
    # Get the tab1 template
    tab1_template = summary_wb['tab1']
    
    # Sort files by date for this client
    client_files_list.sort(key=extract_date_from_filename)
    
    # Process each file for this client
    tabs_created = 0
    for filepath in client_files_list:
        print(f"    Processing file: {os.path.basename(filepath)}")
        
        try:
            # Load the client workbook
            client_wb = load_workbook(filepath)
            
            # Get the first worksheet from the client file
            client_ws = client_wb.active
            
            # Extract date from the client sheet
            sheet_date = extract_date_from_sheet(client_ws)
            
            # If we couldn't find a date in the sheet, use filename date
            if not sheet_date:
                sheet_date = extract_date_from_filename(filepath)
                print(f"      Using filename date: {sheet_date}")
            else:
                print(f"      Found date in sheet: {sheet_date}")
            
            # Create tab name using just the date (since it's a single client file)
            if sheet_date:
                new_tab_name = sheet_date
            else:
                # Fallback to filename without extension
                new_tab_name = os.path.splitext(os.path.basename(filepath))[0]
            
            # Ensure tab name is unique and within Excel limits
            new_tab_name = new_tab_name[:31]  # Excel limit
            counter = 1
            original_name = new_tab_name
            while new_tab_name in summary_wb.sheetnames:
                suffix = f"_{counter}"
                new_tab_name = f"{original_name[:31-len(suffix)]}{suffix}"
                counter += 1
            
            # Create new tab with tab1 format and client data
            print(f"      Creating tab '{new_tab_name}' with tab1 format and client data...")
            new_ws = copy_tab1_format_with_client_data(summary_wb, tab1_template, client_ws, new_tab_name)
            
            print(f"      ✓ Created tab '{new_tab_name}'")
            tabs_created += 1
            
        except Exception as e:
            print(f"      ✗ Error processing {filepath}: {e}")
            continue
    
    # Save the workbook (this will break remaining external links)
    print(f"  💾 Saving workbook without external links...")
    summary_wb.save(output_filename)
    print(f"  ✅ Successfully created {output_filename} with {tabs_created} data tabs")
    return True

def create_summary_files():
    """Main function to create summary Excel files for each client"""
    print("🚀 Starting portfolio data integration - creating individual client summaries...")
    
    # Check if Template.xlsx exists
    if not os.path.exists('Template.xlsx'):
        print("❌ Template.xlsx not found in current directory")
        return
    
    # Read account list to get all clients
    clients_from_account_list = read_account_list()
    if not clients_from_account_list:
        print("⚠️  No clients found in account list, continuing with files found in portfolio...")
    
    # Get client files
    client_files = get_client_files()
    if not client_files:
        print("❌ No client files found")
        return
    
    # Organize client files by client name
    client_data = {}
    for filepath in client_files:
        client_name = extract_client_name(filepath)
        if client_name not in client_data:
            client_data[client_name] = []
        client_data[client_name].append(filepath)
    
    # Filter client_data to only include clients from account_list if available
    if clients_from_account_list:
        filtered_client_data = {}
        for client_name, files in client_data.items():
            if client_name in clients_from_account_list:
                filtered_client_data[client_name] = files
            else:
                print(f"⏭️  Skipping {client_name} - not found in account list")
        client_data = filtered_client_data
    
    if not client_data:
        print("❌ No matching client files found for clients in account list")
        return
    
    # Create summary file for each client
    successful_clients = []
    failed_clients = []
    
    for client_name, client_files_list in client_data.items():
        success = create_summary_for_client(client_name, client_files_list)
        if success:
            successful_clients.append(client_name)
        else:
            failed_clients.append(client_name)
    
    # Final summary
    print(f"\n📊 SUMMARY:")
    print(f"✅ Successfully created summary files for {len(successful_clients)} clients:")
    for client in successful_clients:
        print(f"   - {client}_summary.xlsx")
    
    if failed_clients:
        print(f"❌ Failed to create summary files for {len(failed_clients)} clients:")
        for client in failed_clients:
            print(f"   - {client}")
    
    print(f"\n🔒 All external links have been removed to prevent security warnings.")

if __name__ == "__main__":
    create_summary_files()
